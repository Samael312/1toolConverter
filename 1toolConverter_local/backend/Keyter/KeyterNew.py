#!/usr/bin/env python3
"""
Backend Keyter: procesamiento de archivos Excel -> DataFrame -> Excel procesado.
Procesa múltiples hojas y genera archivo Excel final con l10n multilingüe.
Uso:
    python main.py archivo.xlsx
Genera:
    processed_archivo.xlsx
"""

from typing import List, Optional, Dict, Any
import sys
import os
import logging
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import json

# =====================================================
# CONFIGURACIÓN DE LOGGING
# =====================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# =====================================================
# CONSTANTES
# =====================================================

LIBRARY_COLUMNS = [
    "id", "register", "name", "description", "system_category", "category", "view",
    "sampling", "read", "write", "minvalue", "maxvalue", "unit", "offset",
    "addition", "mask", "value", "length", "general_icon", "alarm", "metadata",
    "l10n", "tags", "type", "parameter_write_byte_position", "mqtt", "json",
    "current_value", "current_error_status", "notes"
]

COLUMN_MAPPING1 = {
    "MODBUS ADDRESS\nCAREL": "register",
    "Name of the variable CAREL": "name",
    "Descripción (ES)": "description",
    "Description (EN)": "description_en",
    "Beschreibung (DE)": "description_de",
    "Description (FR)": "description_fr",
    "Descrizione (IT)": "description_it",
    "Unit": "unit",
    "Min": "minvalue",
    "Max": "maxvalue",
    "Category": "system_category"
}

# =====================================================
# LECTURA DE HOJA
# =====================================================

from io import BytesIO
from typing import Optional, List
import pandas as pd
from openpyxl import load_workbook

def process_excel(
    excel_source: Any,
    sheet_name: Optional[str] = None,
    header_row: int = 5
) -> Optional[pd.DataFrame]:
    """
    Procesa un Excel desde bytes o desde ruta de archivo.
    Si se pasa una ruta, se procesará una sola hoja.
    Si se pasa bytes, se procesarán todas las hojas.
    """
    try:
        if isinstance(excel_source, bytes):
            excel_io = BytesIO(excel_source)
            wb = load_workbook(excel_io, data_only=True, read_only=False, keep_vba=True)
        else:
            wb = load_workbook(excel_source, data_only=True, read_only=False, keep_vba=True)
    except Exception as e:
        logger.error(f"Error al cargar Excel: {e}", exc_info=True)
        return None

    # Si se especifica una hoja, solo esa
    sheetnames = [sheet_name] if sheet_name else wb.sheetnames

    all_dfs = []
    for name in sheetnames:
        logger.info(f"🟢 Procesando hoja '{name}'...")
        ws = wb[name]
        header_cells = list(ws[header_row])
        headers = [str(c.value).strip() if c.value else f"col_{i}" for i, c in enumerate(header_cells)]
        rows = []

        for row in ws.iter_rows(min_row=header_row + 2, max_col=len(header_cells)):
            values = [cell.value for cell in row]
            row_dict = {headers[i]: values[i] for i in range(len(headers))}
            row_dict["sheet_name"] = name
            rows.append(row_dict)

        df_sheet = pd.DataFrame(rows)
        all_dfs.append(df_sheet)

    if not all_dfs:
        return None

    df_total = pd.concat(all_dfs, ignore_index=True)
    df_total = _process_dataframe(df_total)
    df_total = _add_default_columns(df_total)
    df_total = df_total.reindex(columns=LIBRARY_COLUMNS)

    return df_total



# =====================================================
# PROCESAMIENTO DE ARCHIVO COMPLETO
# =====================================================

def process_excel_file(path: str, header_row: int = 5) -> Optional[pd.DataFrame]:
    wb = load_workbook(path, data_only=True)
    logger.info(f"Hojas detectadas: {wb.sheetnames!r}")

    preferred = {"ANALOG", "INTEGER", "BOOL"}
    found = [s for s in wb.sheetnames if s.upper().strip() in preferred]

    if len(found) < 3:
        for s in wb.sheetnames:
            if s not in found and any(k in s.upper() for k in preferred):
                found.append(s)
    if len(found) < 3:
        found = wb.sheetnames[:3]

    logger.info(f"Hojas a procesar: {found!r}")
    processed_dfs = []

    for sheet in found:
        ws = wb[sheet]
        if getattr(ws, "max_row", 0) <= header_row:
            logger.warning(f"Saltando hoja '{sheet}' (vacía o pequeña).")
            continue
        try:
            df_raw = process_excel(path, sheet, header_row)
            if df_raw.empty:
                continue
            df_proc = _process_dataframe(df_raw)
            processed_dfs.append(df_proc)
            logger.info(f"Procesada hoja '{sheet}' ({len(df_proc)} filas).")
        except Exception as e:
            logger.error(f"Error procesando hoja {sheet}: {e}", exc_info=True)

    if not processed_dfs:
        return None

    final_df = pd.concat(processed_dfs, ignore_index=True)
    final_df["id"] = range(1, len(final_df) + 1)
    final_df = _add_default_columns(final_df)
    return final_df.reindex(columns=LIBRARY_COLUMNS)

# =====================================================
# LIMPIEZA Y CLASIFICACIÓN
# =====================================================

def _process_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = df.columns.astype(str).str.strip()
    df = _apply_column_mapping1(df)
    df = _process_specific_columns(df)
    df = _process_access_permissions(df)
    df = _determine_data_length(df)
    df = _apply_deep_classification(df)
    df = _apply_sampling_rules(df)
    df = _apply_view_rules(df)
    df = _apply_localization(df)
    df = _apply_system_configuratiom(df)
    df = _apply_range_rules(df)
    df.dropna(how="all", inplace=True)
    return df

def _apply_column_mapping1(df: pd.DataFrame) -> pd.DataFrame:
    col_map = {}
    # Recorremos columnas tal cual, sin convertirlas a minúsculas.
    for src, dst in COLUMN_MAPPING1.items():
        if src in df.columns:   # sensibilidad a mayúsculas/minúsculas
            col_map[src] = dst

    # Igual para r/w o direction, pero sin convertir todo a minúsculas
    for candidate in ["R/W", "direction"]:
        if candidate in df.columns:
            col_map[candidate] = "access_type"

    df.rename(columns=col_map, inplace=True)
    return df


def _process_specific_columns(df: pd.DataFrame) -> pd.DataFrame:
    if 'offset' in df.columns:
        df['offset'] = pd.to_numeric(
            df['offset'].astype(str).replace(['', '---', 'nan'], None),
            errors='coerce'
        ).fillna(0.0)

    if 'unit' in df.columns:
        df['unit'] = df['unit'].astype(str).replace(['0', '---', None], ' ').str.strip()

    if 'system_category' in df.columns:
        df['system_category'] = df['system_category'].astype(str).str.upper().str.strip()

        mapping = {
            'CONSIGNA': 'SET_POINT',
            'E.ANALOGICA': 'ANALOG_INPUT',
            'ALARMA': 'ALARM',
            'COMANDO': 'COMMAND',
            'E.DIGITAL': 'DIGITAL_INPUT',
            'ESTADO': 'STATUS',
            'PARAMETRO': 'CONFIG_PARAMETER',
            'S.ANALOGICA': 'ANALOG_OUTPUT',
            'S.DIGITAL': 'DIGITAL_OUTPUT',
            'SISTEMA': 'SYSTEM'
        }

        df['system_category'] = df['system_category'].replace(mapping)
        df = df[~df['system_category'].isin(['NONE', '','NAN', None])].copy()
    
        df.loc[df['system_category'] == 'SET_POINT', 'offset'] = 0.1
    
    if 'name' in df.columns:
        df['name'] = df['name'].astype(str).str.upper().str.strip()
        df = df[~df['name'].isin(['NONE', '','NAN', None])].copy()

    for col in ['minvalue', 'maxvalue']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    if 'description' in df.columns:
        df['description'] = df['description'].astype(str).str.slice(0, 60)

    return df

def _apply_range_rules(df: pd.DataFrame) -> pd.DataFrame:
    """Corrige valores fuera del rango [-32767, 32767] y marca los ajustes realizados."""
    
    # --- 0️⃣ Validación general ---
    if "system_category" in df.columns:
        df["system_category"] = df["system_category"].astype(str).str.upper().str.strip()
        system_type = df["system_category"]
    else:
        system_type = pd.Series([], dtype=str)

    # --- 1️⃣ Aplicar corrección de rango ---
    if "value" in df.columns:
        # Guardar valor original (opcional)
        df["original_value"] = df["value"]
        
        # Limitar valores al rango permitido
        df["value"] = df["value"].clip(lower=-32767, upper=32767)
        
        # Indicar si fue corregido
        df["range_check"] = np.where(
            df["original_value"].between(-32767, 32767, inclusive="both"),
            "OK",
            "Valor corregido al límite [-32767 ~ 32767]"
        )
    else:
        df["range_check"] = "Columna 'value' no encontrada"

    return df

def _process_access_permissions(df: pd.DataFrame) -> pd.DataFrame:
    df['read'] = 0
    df['write'] = 0

    df.columns = df.columns.astype(str).str.strip()
    only_read = pd.Series(False, index=df.index)
    rw = pd.Series(False, index=df.index)

    if 'access_type' in df.columns:
        access = df['access_type'].astype(str).str.upper().str.strip()

        only_read = access == 'R'
        rw = access == 'R/W'

        df.loc[access == 'R', 'read'] = 4
        df.loc[access == 'W', 'write'] = 6
        df.loc[access == 'R/W', ['read', 'write']] = [3, 6]
        df.drop(columns=['access_type'], inplace=True, errors='ignore')

    if 'system_category' in df.columns:
        system_type = df['system_category'].astype(str).str.upper().str.strip()
        df.loc[only_read & system_type.isin(['ANALOG_INPUT', 'ANALOG_OUTPUT', 'SYSTEM']), 'read'] = 3
        df.loc[only_read & system_type.isin(['ALARM']), 'read'] = 1
        df.loc[only_read & system_type.isin(['STATUS']), 'read'] = 4
        df.loc[only_read & system_type.isin(['COMMAND']), 'write'] = 6
        df.loc[rw & system_type.isin(['SET_POINT']), ['read', 'write']] = [3, 6]
        df.loc[rw & system_type.isin(['CONFIG_PARAMETER','DIGITAL_OUTPUT']), ['read', 'write']] = [1, 5]
        df.loc[rw & system_type.isin(['DIGITAL_INPUT']), 'read'] = 1

    return df

def _determine_data_length(df: pd.DataFrame) -> pd.DataFrame:
    category_mapping = {
        'SET_POINT': '16bit/s16',
        'ANALOG_INPUT': '16bit/s16',
        'ALARM': '1bit',
        'COMMAND': '1bit',
        'DIGITAL_INPUT': '1bit',
        'STATUS': '1bit',
        'CONFIG_PARAMETER': '16bit/s16',
        'ANALOG_OUTPUT': '16bit/s16',
        'DIGITAL_OUTPUT': '1bit',
        'SYSTEM': '16bit'
    }

    df["length"] = None
    if "system_category" in df.columns:
        df.loc[df["system_category"].isin(category_mapping.keys()), "length"] = df["system_category"].map(category_mapping)

    mask_16bit_s16 = df["length"] == "16bit/s16"
    if "minvalue" in df.columns and "maxvalue" in df.columns:
        df.loc[mask_16bit_s16 & ((df["minvalue"] < 0) | (df["maxvalue"] < 0)), "length"] = "s16"
        df.loc[mask_16bit_s16 & ~((df["minvalue"] < 0) | (df["maxvalue"] < 0)), "length"] = "16bit"

    mask_none = df["length"].isna()
    if "minvalue" in df.columns and "maxvalue" in df.columns:
        df.loc[mask_none, "length"] = "16bit"
        df.loc[mask_none & ((df["minvalue"] < 0) | (df["maxvalue"] < 0)), "length"] = "s16"

    return df

def _apply_deep_classification(df: pd.DataFrame) -> pd.DataFrame:
    
    if "system_category" not in df.columns:
        df["system_category"] = "STATUS"
    df["system_category"] = df["system_category"].astype(str).str.upper().str.strip()
    df["system_category"] = np.where(df["system_category"].eq("ALARM"), "ALARM", df["system_category"])

  
   
    return df

def _apply_system_configuratiom(df:pd.DataFrame) -> pd.DataFrame:
    if "system_category" in df.columns:
        system_type = df['system_category'].astype(str).str.upper().str.strip()
        df["tags"] = np.where(system_type.isin(["SYSTEM"]), '["library_identifier"]', "[]")
    
    if "name" in df.columns:
        # Eliminar espacios y asegurar tipo string
        df["name"] = df["name"].astype(str).str.strip()
        
        # Contar repeticiones
        duplicates = df.groupby("name").cumcount()
        
        # Agregar sufijo solo a duplicados (primer valor tiene índice 0)
        df["name"] = np.where(
            duplicates == 0,
            df["name"],
            df["name"] + "_" + (duplicates + 1).astype(str)
        )
    
    return df 

def _apply_sampling_rules(df: pd.DataFrame) -> pd.DataFrame:
    mapping = {"ALARM": 30, "SET_POINT": 300, "DEFAULT": 0, "COMMAND": 0, "STATUS": 60, "SYSTEM": 0, "CONFIG_PARAMETER":0}
    df["sampling"] = df["system_category"].map(mapping).fillna(0)
    return df

def _apply_view_rules(df: pd.DataFrame) -> pd.DataFrame:
    df.loc[df['system_category'] == 'STATUS', 'view'] = 'basic'
    df.loc[df['system_category'] == 'ALARM', 'view'] = 'simple'
    view = {
        'ALARM': 'simple',
        'SET_POINT': "simple",
        'DEFAULT': "simple",
        'COMMAND': "simple",
        'STATUS': "basic",
        'CONFIG_PARAMETER': "simple",
        'SYSTEM': 'simple'
    }
    df['view'] = df['system_category'].map(view).fillna('simple')
    return df

def _apply_localization(df: pd.DataFrame) -> pd.DataFrame:
    """Crea campo l10n multilingüe a partir de columnas de descripción."""

    langs = {
        "description_en": "en_US",
        "description": "es_ES",
        "description_de": "de_DE",
        "description_fr": "fr_FR",
        "description_it": "it_IT"
    }

    def build_l10n(row):
        translations = {}
        for col, lang in langs.items():
            if col in row and pd.notna(row[col]):
                # Convertir a string y cortar a 60 caracteres
                desc = str(row[col]).strip()[:60]
                translations[lang] = {
                    "name": None,
                    "_type": "languages",
                    "category": None,
                    "description": desc
                }

        # Si no hay traducciones, incluir al menos es_ES vacío
        if not translations:
            translations = {
                "es_ES": {
                    "name": None,
                    "_type": "languages",
                    "category": None,
                    "description": None
                }
            }
        
        l10n_dict = {
            "_type": "l10n",
            "default_lang": "es_ES",
            "translations": translations
        }

        # Convertir a JSON con formato sin espacios innecesarios
        return json.dumps(l10n_dict, ensure_ascii=False, separators=(',', ':'))

    df["l10n"] = df.apply(build_l10n, axis=1)
    return df

def _add_default_columns(df: pd.DataFrame) -> pd.DataFrame:
    defaults = {
        "addition": 0,
        "mask": 0,
        "value": 0,
        "alarm": '{"severity":"none"}',
        "metadata": "[]",
        "tags": "[]",
        "type": "modbus"
    }
    for c, v in defaults.items():
        if c not in df.columns:
            df[c] = v
    return df

# =====================================================
# MAIN
# =====================================================

def _output_path(path: str) -> str:
    base = os.path.basename(path)
    name, _ = os.path.splitext(base)
    return os.path.join(os.path.dirname(path), f"processed_{name}.xlsx")

def main(argv: List[str]):
    if len(argv) < 2:
        print("Uso: python main.py archivo.xlsx")
        sys.exit(1)

    input_path = argv[1]
    output_path = _output_path(input_path)

    df = process_excel_file(input_path)
    if df is None:
        logger.error("No se generó DataFrame final.")
        sys.exit(2)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="processed")

    logger.info(f"Archivo exportado correctamente: {output_path}")
    print(f"✅ Archivo generado: {output_path}")

if __name__ == "__main__":
    main(sys.argv)
